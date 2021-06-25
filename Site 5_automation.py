from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from random_user_agent.user_agent import UserAgent


from random_user_agent.params import SoftwareName, OperatingSystem
import openpyxl
import random
import time

search_char = ['L','S','T']

software_names = [SoftwareName.CHROME.value]
operating_systems = [OperatingSystem.WINDOWS.value,
                     OperatingSystem.LINUX.value]
user_agent_rotator = UserAgent(software_names = software_names,
                       operating_systems = operating_systems,
                       limit=100)
user_agent = user_agent_rotator.get_random_user_agent()




def click_element(elem_type=0, elem_selector=0):
    if(elem_type == 'css'):
        elem = browser.find_element_by_css_selector(elem_selector)
    elem.click

def sleep_fuc():
    time.sleep(random.randrange(2,4))
    
def get_search_window(srch_address,srch_prim_plan,srch_sub_plan,row_num,col_num,active_sheet):
    global browser
    global software_names, operating_systems, user_agent_rotator, user_agent

    chrome_options = Options()
    chrome_options.add_argument("user-agent ={}".format(user_agent))

    browser = webdriver.Chrome(options=chrome_options)
    browser.get('https://www.blueshieldca.com/fap/app/find-a-doctor.html')
    
    

    browser.switch_to.window(browser.window_handles[1])

    doctor_button=browser.find_element_by_id('Doctor')

    doctor_button.click()

    #guest_button=find_element_by_id('nonMember')
    WebDriverWait(browser,2).until(EC.element_to_be_clickable((By.ID,'nonMember'))).click()

    #guest_button.click()

    location_input=browser.find_element_by_id('autocomplete')

    #location_input.click()
    #location_input.send_keys('Los Angeles, CA 90002, USA')



    sleep_fuc()

    while(browser.current_url != 'https://www.blueshieldca.com/fad/plans/chooseplan'):
        location_input.clear()
        location_input.send_keys(srch_address)
        
        location_input.send_keys(Keys.LEFT)
        location_input.send_keys(Keys.RIGHT)
        
        action = webdriver.common.action_chains.ActionChains(browser)
        action.move_to_element_with_offset(location_input, 100,60)
        action.click()
        action.perform()

        sleep_fuc()

        #CLICK CONTINUE
        #click_element('css','#main > div > app-location > div.location-container > div > div.wrapper-location.col-sm-12.col-md-10.col-lg-7 > button')
        continue_button = browser.find_element_by_css_selector('#main > div > app-location > div.location-container > div > div.wrapper-location.col-sm-12.col-md-10.col-lg-7 > button')
        sleep_fuc()

        continue_button.click()


    sleep_fuc()
    for tries in range(0,5):
        try:
            browser.refresh()
            allHTML = browser.find_element_by_tag_name('html')
            allHTML.send_keys(Keys.PAGE_DOWN)
            time.sleep(1)
            primary_plan_button = browser.find_element_by_css_selector('#main > div > app-choose-plans > div > div > div > div > div > button.btn.btn-primary.planButtonLeft')
            primary_plan_button.click()

            #CLICK PRIMARY PLAN DROPDOWN
            #click_element('css','#primaryPlanDropdown > div')
            #htmlElem = browser.find_element_by_tag_name('html')
            #htmlElem.send_keys(Keys.PAGE_DOWN)
            primary_plan_dropdown = browser.find_element_by_xpath('//*[@id="primaryPlanDropdown"]')
            browser.execute_script("arguments[0].click();", primary_plan_dropdown)
            time.sleep(2)


            primary_plan_ul = browser.find_element_by_xpath('//*[@id="main"]/div/app-plans/div[1]/div/div/div/form/div/div/div/div[3]/div[2]/ul')
            primary_plans = primary_plan_ul.find_elements_by_tag_name('li')

            srch_prim_plan = srch_prim_plan.replace(" ","").lower()
            for plan in primary_plans:
                plan_text = plan.text
                plan_text = plan_text.replace(" ","").lower()
                if (plan_text == srch_prim_plan):
                    plan.click()
                    break

            try:
                sub_plan_dropdown = browser.find_element_by_xpath('//*[@id="subPlanDropdown"]')
                browser.execute_script("arguments[0].click();", sub_plan_dropdown)

                allHTML = browser.find_element_by_tag_name('html')
                allHTML.send_keys(Keys.PAGE_DOWN)
                time.sleep(1)

                sub_plan_ul = browser.find_element_by_xpath('//*[@id="main"]/div/app-plans/div[1]/div/div/div/form/div/div/div/div[3]/div[3]/ul')
                sub_plans = sub_plan_ul.find_elements_by_tag_name('li')

                #Create search pattern
            
                srch_sub_plan = srch_sub_plan.replace(" ","")
                
                for sub in sub_plans:
                    sub_text = sub.text
                    sub_text = sub_text.replace(" ","")
                    
                    if (sub_text == srch_sub_plan):
                        sub.click()
                        break
                
            except:
                print('No Subplan')
        

        
                
            sleep_fuc()
            plan_continue = browser.find_element_by_xpath('//*[@id="continuePlan"]')
            browser.execute_script("arguments[0].click();", plan_continue)
            sleep_fuc()
            return True
        except Exception as e:
            #print(e)
            print('Request Denied, try# '+str(tries))
            #browser.quit()
            #software_names = [SoftwareName.CHROME.value]
            #operating_systems = [OperatingSystem.WINDOWS.value,
                                 #OperatingSystem.LINUX.value]
            #user_agent_rotator = UserAgent(software_names = software_names,
                                   #operating_systems = operating_systems,
                                   #limit=100)
            #user_agent = user_agent_rotator.get_random_user_agent()
            browser.back()
            time.sleep(2*tries)
            #get_search_window(srch_address,srch_prim_plan,srch_sub_plan,row_num,col_num,active_sheet)
            #active_sheet.cell(row = row_num, column = col_num).value = "0 (Manually Check)"
    return False
#####################

def get_doctor_count(row_num,col_num,active_sheet):
    global search_char
    #doctor_name = browser.find_element_by_xpath('//*[@id="Enter last name"]/label')
    #doctor_name.click()

    WebDriverWait(browser,2).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="Enter last name"]/label'))).click()
    time.sleep(0.3)

    enter_name = WebDriverWait(browser,2).until(EC.presence_of_element_located((By.XPATH,'//*[@id="provider_name"]')))
    enter_name.clear()
    enter_name.send_keys(search_char[0])

    search_btn = browser.find_element_by_xpath('//*[@id="searchBtn"]')
    search_btn.click()

    sleep_fuc()

    try:
        miles = browser.find_element_by_xpath('//*[@id="searchResultMsg"]/a')
        browser.execute_script("arguments[0].click();", miles)
    except:
        total_doctors = "0 (Failed at miles)"
        active_sheet.cell(row = row_num, column = col_num).value = total_doctors
        print('Failed at miles')
        return

    miles_dropdown = browser.find_element_by_xpath('//*[@id="dropdownRadiusFilter"]')
    browser.execute_script("arguments[0].click();", miles_dropdown)

    time.sleep(1.5)

    #miles_ul = browser.find_element_by_xpath('//*[@id="affix_el"]/app-filtersort/div/div[2]/form/fieldset/div[1]/div[1]/div/ul')
    miles_100 = browser.find_element_by_xpath('//*[@id="select_lo7"]')
    miles_100.click()

    apply_btn = browser.find_element_by_xpath('//*[@id="filter-button-apply"]')
    apply_btn.click()
    time.sleep(4)
    
    try:
        total_doctors = browser.find_element_by_xpath('//*[@id="searchResultCount"]/span[1]')
    except:
        total_doctors = "0"

    active_sheet.cell(row = row_num, column = col_num).value = total_doctors.text
    print('Success!: '+ str(row_num)+', '+str(col_num))

    for index in range(1,3):
        col_num = col_num + 1
        enter_name = WebDriverWait(browser,2).until(EC.presence_of_element_located((By.XPATH,'//*[@id="provider_name"]')))
        enter_name.clear()
        enter_name.send_keys(search_char[index])


        WebDriverWait(browser,2).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="searchBtn"]'))).click()
        time.sleep(5)
        try:
            total_doctors = browser.find_element_by_xpath('//*[@id="searchResultCount"]/span[1]')
        except:
            total_doctors = "0"
    
        active_sheet.cell(row = row_num, column = col_num).value = total_doctors.text
        print('Success!: '+ str(row_num)+', '+str(col_num))

####################



def main():
    #OPEN EXCEL FILE
    #print('Please input the whole file name. (must be .xlsx file)')
    file = 'Site 5 Product Research Jun21.xlsx'
    wb = openpyxl.load_workbook(file)


    print('Start on which row?')
    start_row = input()
    print('Start on which column?')
    start_column = input()
    
    sheet = wb['Sheet1']

    for row in range(start_row, sheet.max_row +1):
        sub_plan = sheet['D' + str(row)].value
        prim_plan = sheet['E' + str(row)].value

        
        

        for col in range(start_column, sheet.max_column + 1, 3):
            search_value = sheet.cell(row = 1, column = col).value
            zip_code = search_value[0:5]
            #search_char = search_value[-1]

            sheet = wb['Sheet2']
            for zip_row in range(1,sheet.max_row+1):
                if (str(zip_code) == str(sheet['A'+str(zip_row)].value)):
                    address = sheet['B'+str(zip_row)].value
                    break
                    
            sheet = wb['Sheet1']
            proceed = get_search_window(address,prim_plan,sub_plan,row,col,sheet)
            if (proceed):
                get_doctor_count(row,col,sheet)
            wb.save(file)
            browser.quit()

        start_column = 7


    
    print('Done!')

main()
